from io import BytesIO
from decimal import Decimal

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.core.database import get_db
from app.models.student import Student

from app.models.user import User



router = APIRouter(
    prefix="/imports",
    tags=["Excel Import"],
)


COLUMN_ALIASES = {
    "matric number": "matric_number",
    "matriculation number": "matric_number",

    "name of student": "full_name",
    "student name": "full_name",
    "name": "full_name",

    "dress": "dressing_appearance",
    "dressing": "dressing_appearance",
    "dressing & appearance": "dressing_appearance",

    "oral presentation": "oral_presentation",

    "slide presentation": "slide_presentation",

    "depth of understanding": "depth_of_understanding",

    "project implementation": "project_implementation",

    "referencing & documentation": "referencing_documentation",
    "referencing and documentation": "referencing_documentation",

    "contribution & originality": "contribution_originality",
    "contribution and originality": "contribution_originality",

    "professional conduct": "professional_conduct",
}


def clean_header(value):
    if value is None:
        return ""

    return " ".join(str(value).strip().lower().split())


def get_recommendation(total):
    if total >= 50:
        return "Pass"

    return "Fail"


@router.post("/excel")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):

    if not file.filename:
        raise HTTPException(
            status_code=400,
            detail="No file selected.",
        )

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx Excel files are supported.",
        )

    contents = await file.read()

    try:
        workbook = load_workbook(
            filename=BytesIO(contents),
            data_only=True,
        )
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Unable to read the Excel file.",
        )

    imported = 0
    skipped = 0
    errors = []

    for sheet_name in workbook.sheetnames:

        # Ignore index/summary sheets
        if sheet_name.strip().lower() in {
            "index",
            "summary",
            "contents",
        }:
            continue

        ws = workbook[sheet_name]

        # Headers are expected around row 8.
        # Search first 12 rows so lecturers do not have to
        # follow one exact Excel layout.
        header_row = None
        headers = {}

        for row_number in range(1, min(ws.max_row, 12) + 1):

            candidate_headers = {}

            for column in range(1, ws.max_column + 1):

                value = clean_header(
                    ws.cell(row_number, column).value
                )

                if value:
                    candidate_headers[value] = column

            if (
                "matric number" in candidate_headers
                or "matriculation number" in candidate_headers
            ):
                header_row = row_number
                headers = candidate_headers
                break

        if header_row is None:
            continue

        matric_column = (
            headers.get("matric number")
            or headers.get("matriculation number")
        )

        for row_number in range(
            header_row + 1,
            ws.max_row + 1,
        ):

            matric_value = ws.cell(
                row_number,
                matric_column,
            ).value

            # Empty matric number means:
            # no student record on this row.
            if (
                matric_value is None
                or str(matric_value).strip() == ""
            ):
                skipped += 1
                continue

            matric = str(matric_value).strip()

            try:

                student = (
                    db.query(Student)
                    .filter(
                        Student.matric_number == matric,
                        Student.is_deleted == False,
                    )
                    .first()
                )

                if not student:
                    errors.append(
                        f"{sheet_name}, row {row_number}: "
                        f"student {matric} not found."
                    )
                    continue

                scores = {}

                for excel_header, ems_field in COLUMN_ALIASES.items():

                    if excel_header not in headers:
                        continue

                    if ems_field in {
                        "matric_number",
                        "full_name",
                    }:
                        continue

                    value = ws.cell(
                        row_number,
                        headers[excel_header],
                    ).value

                    if value is None or str(value).strip() == "":
                        continue

                    try:
                        scores[ems_field] = Decimal(str(value))
                    except Exception:
                        errors.append(
                            f"{sheet_name}, row {row_number}: "
                            f"invalid score '{value}' "
                            f"for {excel_header}."
                        )

                # Need all eight assessment scores
                required_scores = [
                    "dressing_appearance",
                    "oral_presentation",
                    "slide_presentation",
                    "depth_of_understanding",
                    "project_implementation",
                    "referencing_documentation",
                    "contribution_originality",
                    "professional_conduct",
                ]

                missing = [
                    field
                    for field in required_scores
                    if field not in scores
                ]

                if missing:
                    errors.append(
                        f"{sheet_name}, row {row_number}: "
                        f"missing scores: {', '.join(missing)}."
                    )
                    continue

                total_score = sum(
                    scores[field]
                    for field in required_scores
                )

                # Database validation
                if total_score < 0 or total_score > 100:
                    errors.append(
                        f"{sheet_name}, row {row_number}: "
                        f"total score {total_score} is outside 0-100."
                    )
                    continue

                assessment = Assessment(
                    student_id=student.id,
                    assessor_id=current_user.id,

                    dressing_appearance=scores[
                        "dressing_appearance"
                    ],

                    oral_presentation=scores[
                        "oral_presentation"
                    ],

                    slide_presentation=scores[
                        "slide_presentation"
                    ],

                    depth_of_understanding=scores[
                        "depth_of_understanding"
                    ],

                    project_implementation=scores[
                        "project_implementation"
                    ],

                    referencing_documentation=scores[
                        "referencing_documentation"
                    ],

                    contribution_originality=scores[
                        "contribution_originality"
                    ],

                    professional_conduct=scores[
                        "professional_conduct"
                    ],

                    total_score=total_score,

                    recommendation=get_recommendation(
                        total_score
                    ),

                    remarks=None,

                    is_deleted=False,
                )

                db.add(assessment)
                imported += 1

            except Exception as exc:

                errors.append(
                    f"{sheet_name}, row {row_number}: "
                    f"{str(exc)}"
                )

    try:
        db.commit()

    except Exception as exc:
        db.rollback()

        raise HTTPException(
            status_code=500,
            detail=f"Excel import failed: {str(exc)}",
        )

    return {
        "message": "Excel import completed.",
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
    }
