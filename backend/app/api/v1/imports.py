import re
from io import BytesIO
from decimal import Decimal, InvalidOperation

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    UploadFile,
    status,
)
from pydantic import ValidationError
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.core.database import get_db
from app.core.auth_dependency import get_current_user

from app.models.student import Student
from app.schemas.assessment import AssessmentCreate
from app.services.assessment_service import create_or_update_assessment


router = APIRouter(
    prefix="/imports",
    tags=["Excel Import"],
)


# Header text is matched after stripping a trailing "(10)"-style weight
# annotation, so this file works whether or not the sheet includes the
# point values in the header.
COLUMN_ALIASES = {
    "matric number": "matric_number",
    "matriculation number": "matric_number",

    "name of student": "full_name",
    "student name": "full_name",
    "name": "full_name",

    "dress": "dress",
    "dressing": "dress",
    "dressing & appearance": "dress",

    "report format": "report_format",
    "format of report": "report_format",

    "problem solved": "problem_solved",

    "clarity of write-up": "clarity_of_writeup",
    "clarity of writeup": "clarity_of_writeup",
    "clarity of write up": "clarity_of_writeup",

    "result presentation": "result_presentation",

    "evidence of understanding": "evidence_of_understanding",

    "knowledge contribution": "knowledge_contribution",

    "reference": "reference",
    "references": "reference",
}


REQUIRED_SCORES = [
    "dress",
    "report_format",
    "problem_solved",
    "clarity_of_writeup",
    "result_presentation",
    "evidence_of_understanding",
    "knowledge_contribution",
    "reference",
]


def clean_header(value):
    if value is None:
        return ""

    text = str(value).replace("\n", " ").strip().lower()

    # Strip a trailing weight annotation like "(10)" or "(15)"
    text = re.sub(r"\(\s*\d+\s*\)\s*$", "", text).strip()

    return " ".join(text.split())


def find_headers(ws):
    for row_number in range(1, min(ws.max_row, 12) + 1):
        candidate_headers = {}

        for column in range(1, ws.max_column + 1):
            value = clean_header(ws.cell(row_number, column).value)

            if value:
                candidate_headers[value] = column

        if (
            "matric number" in candidate_headers
            or "matriculation number" in candidate_headers
        ):
            return row_number, candidate_headers

    return None, {}


def parse_score(value):
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()

        if not value:
            return None

    try:
        return Decimal(str(value))

    except (InvalidOperation, ValueError, TypeError):
        return None


@router.post("/excel")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file selected.",
        )

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx Excel files are supported.",
        )

    contents = await file.read()

    if not contents:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded Excel file is empty.",
        )

    try:
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)

    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read the Excel file.",
        )

    imported = 0
    skipped = 0
    errors = []

    for sheet_name in workbook.sheetnames:

        if sheet_name.strip().lower() in {"index", "summary", "contents"}:
            continue

        ws = workbook[sheet_name]

        header_row, headers = find_headers(ws)

        if header_row is None:
            skipped += 1
            continue

        matric_column = headers.get("matric number") or headers.get(
            "matriculation number"
        )

        if matric_column is None:
            skipped += 1
            continue

        for row_number in range(header_row + 1, ws.max_row + 1):

            matric_value = ws.cell(row_number, matric_column).value

            if matric_value is None or str(matric_value).strip() == "":
                skipped += 1
                continue

            matric = str(matric_value).strip()

            try:
                student = (
                    db.query(Student)
                    .filter(
                        Student.matric_number == matric,
                        Student.is_deleted == False,
                    )
                    .first()
                )

                if not student:
                    errors.append(
                        f"{sheet_name}, row {row_number}: "
                        f"student {matric} not found in EMS."
                    )
                    continue

                scores = {}
                invalid_scores = []

                for excel_header, ems_field in COLUMN_ALIASES.items():

                    if excel_header not in headers:
                        continue

                    if ems_field in {"matric_number", "full_name"}:
                        continue

                    value = ws.cell(row_number, headers[excel_header]).value

                    if value is None or str(value).strip() == "":
                        continue

                    score = parse_score(value)

                    if score is None:
                        invalid_scores.append(f"{excel_header}='{value}'")
                        continue

                    scores[ems_field] = score

                if invalid_scores:
                    errors.append(
                        f"{sheet_name}, row {row_number}: "
                        f"invalid scores: {', '.join(invalid_scores)}."
                    )
                    continue

                missing = [
                    field for field in REQUIRED_SCORES if field not in scores
                ]

                if missing:
                    errors.append(
                        f"{sheet_name}, row {row_number}: "
                        f"missing scores: {', '.join(missing)}."
                    )
                    continue

                assessment_data = AssessmentCreate(
                    student_id=student.id,
                    remarks=None,
                    **scores,
                )

                create_or_update_assessment(
                    db=db,
                    assessment_data=assessment_data,
                    assessor_id=current_user.id,
                    assessment_type="internal",
                )

                imported += 1

            except ValidationError as exc:
                errors.append(
                    f"{sheet_name}, row {row_number}: "
                    f"score out of range: {exc}"
                )

            except Exception as exc:
                db.rollback()

                errors.append(
                    f"{sheet_name}, row {row_number}: {str(exc)}"
                )

    return {
        "message": "Excel import completed.",
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
    }
